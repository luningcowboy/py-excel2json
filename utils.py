from openpyxl import load_workbook
from openpyxl import Workbook

def excel2Json(filePath, mainKeyIdx):
    wb = load_workbook(filePath)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    rows = ws.rows
    cols = ws.columns
    keys = []
    jsData = {}
    for i, row in enumerate(rows):
        # 生成key值
        if i == 0:
            keys = []
            keys = [col.value for col in row]
            continue

        line = [col.value for col in row]
        obj = {}
        for j in range(len(line)):
            obj[keys[j]] = line[j]
        jsData[line[mainKeyIdx]] = line
    return jsData

def excelFilter(keyFile, resFile, keyCol, resCol, outputFile):
    keyFileData = excel2Json(keyFile, keyCol)
    resFileData = excel2Json(resFile, resCol)
    outPutData = []
    failedKeys = []
    for key in keyFileData:
        if key in resFileData.keys():
            outPutData.append(resFileData[key])
        else:
            # 深度检查
            finded = False
            findedKeys = []
            for k2 in resFileData.keys():
                if str(key) in k2:
                    outPutData.append(resFileData[k2])
                    findedKeys.append(k2)
                    finded = True
            if not finded:
                failedKeys.append(str(key))
            else:
                print("findedKeys:", str(key), findedKeys)


    # 写入xlsx
    wb = Workbook()
    ws = wb.create_sheet()
    rowIdx = 1
    colIdx = 1
    for data in outPutData:
        colIdx = 1
        for v in data:
            ws.cell(row = rowIdx, column = colIdx, value = v)
            colIdx = colIdx + 1
        rowIdx = rowIdx + 1
    wb.save(filename='./test.xlsx')
    print(failedKeys, len(failedKeys))


